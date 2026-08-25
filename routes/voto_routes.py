from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash
from flask_login import login_required, current_user
from database import db, peru_now
from models.personero import Colegio, Mesa, Personero
from models.voto import Voto, VotoEspecial, PARTIDOS, CARGOS
from models.activity import ActivityLog

votos_bp = Blueprint("votos", __name__)


@votos_bp.route("/registrar-votos", methods=["GET", "POST"])
@login_required
def registrar_votos():
    colegios = Colegio.query.filter_by(is_active=True).order_by(Colegio.nombre).all()

    if request.method == "POST":
        colegio_id = request.form.get("colegio_id", type=int)
        mesa_id = request.form.get("mesa_id", type=int)

        if not colegio_id or not mesa_id:
            flash("Seleccione colegio y mesa.", "error")
            return redirect(url_for("votos.registrar_votos"))

        for partido in PARTIDOS:
            key = f"voto_{partido['id']}"
            valor = request.form.get(key, 0, type=int)
            if valor > 0:
                existing = Voto.query.filter_by(
                    mesa_id=mesa_id, partido_id=partido["id"]
                ).first()
                if existing:
                    existing.votos = valor
                else:
                    db.session.add(Voto(
                        mesa_id=mesa_id,
                        colegio_id=colegio_id,
                        partido_id=partido["id"],
                        partido_nombre=partido["nombre"],
                        partido_sigla=partido["sigla"],
                        votos=valor,
                        registrado_por=current_user.id,
                    ))

        tipos_especiales = ["BLANCO", "NULO", "IMPUGNADO"]
        for tipo in tipos_especiales:
            key = f"voto_{tipo.lower()}"
            valor = request.form.get(key, 0, type=int)
            if valor >= 0:
                existing = VotoEspecial.query.filter_by(
                    mesa_id=mesa_id, tipo=tipo
                ).first()
                if existing:
                    existing.cantidad = valor
                else:
                    db.session.add(VotoEspecial(
                        mesa_id=mesa_id,
                        colegio_id=colegio_id,
                        tipo=tipo,
                        cantidad=valor,
                        registrado_por=current_user.id,
                    ))

        entry = ActivityLog(
            user_id=current_user.id,
            tipo="VOTOS_REGISTRADOS_ADMIN",
            detalle=f"Registro manual de votos - Mesa {mesa_id}",
            ip_address=request.remote_addr,
            user_agent=request.headers.get("User-Agent", "")[:300],
        )
        db.session.add(entry)
        db.session.commit()
        flash("Votos registrados exitosamente.", "success")
        return redirect(url_for("votos.registrar_votos"))

    return render_template("registrar_votos.html", colegios=colegios, partidos=PARTIDOS)


@votos_bp.route("/api/mesas/<int:colegio_id>")
@login_required
def api_mesas(colegio_id):
    mesas = Mesa.query.filter_by(colegio_id=colegio_id, is_active=True).order_by(Mesa.numero).all()
    return jsonify([m.to_dict() for m in mesas])


@votos_bp.route("/api/votos-resumen")
@login_required
def api_votos_resumen():
    colegio_id = request.args.get("colegio_id", type=int)
    mesa_id = request.args.get("mesa_id", type=int)

    query = Voto.query
    query_especial = VotoEspecial.query

    if colegio_id:
        query = query.filter_by(colegio_id=colegio_id)
        query_especial = query_especial.filter_by(colegio_id=colegio_id)
    if mesa_id:
        query = query.filter_by(mesa_id=mesa_id)
        query_especial = query_especial.filter_by(mesa_id=mesa_id)

    votos_por_partido = {}
    for v in query.all():
        if v.partido_nombre not in votos_por_partido:
            votos_por_partido[v.partido_nombre] = {"sigla": v.partido_sigla, "votos": 0, "id": v.partido_id}
        votos_por_partido[v.partido_nombre]["votos"] += v.votos

    votos_especiales = {}
    total_especial = 0
    for ve in query_especial.all():
        votos_especiales[ve.tipo] = ve.cantidad
        total_especial += ve.cantidad

    total_partidos = sum(v["votos"] for v in votos_por_partido.values())
    total_general = total_partidos + total_especial

    return jsonify({
        "por_partido": votos_por_partido,
        "especiales": votos_especiales,
        "total_partidos": total_partidos,
        "total_especial": total_especial,
        "total_general": total_general,
    })


@votos_bp.route("/api/exportar/excel")
@login_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import io

    colegio_id = request.args.get("colegio_id", type=int)

    query_votos = Voto.query
    query_especial = VotoEspecial.query

    if colegio_id:
        query_votos = query_votos.filter_by(colegio_id=colegio_id)
        query_especial = query_especial.filter_by(colegio_id=colegio_id)

    votos = query_votos.order_by(Voto.cargo, Voto.partido_nombre).all()
    especiales = query_especial.order_by(VotoEspecial.cargo).all()

    wb = Workbook()

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    number_format = '#,##0'

    ws = wb.active
    ws.title = "Resumen por Partido"

    ws.merge_cells("A1:E1")
    ws["A1"] = "REPORTE DE VOTOS - ELECCIONES 2026"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="4F46E5")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generado: {peru_now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Partido", "Sigla", "Cargo", "Mesa", "Votos"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row = 5
    for v in votos:
        ws.cell(row=row, column=1, value=v.partido_nombre).border = thin_border
        ws.cell(row=row, column=2, value=v.partido_sigla).border = thin_border
        ws.cell(row=row, column=3, value=v.cargo).border = thin_border
        ws.cell(row=row, column=4, value=v.mesa.numero if v.mesa else 0).border = thin_border
        c = ws.cell(row=row, column=5, value=v.votos)
        c.number_format = number_format
        c.border = thin_border
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12

    ws2 = wb.create_sheet("Votos Especiales")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "VOTOS ESPECIALES"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="4F46E5")
    ws2["A1"].alignment = Alignment(horizontal="center")

    headers2 = ["Tipo", "Cargo", "Mesa", "Cantidad"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = thin_border

    row = 4
    for e in especiales:
        ws2.cell(row=row, column=1, value=e.tipo).border = thin_border
        ws2.cell(row=row, column=2, value=e.cargo).border = thin_border
        ws2.cell(row=row, column=3, value=e.mesa.numero if e.mesa else 0).border = thin_border
        c = ws2.cell(row=row, column=4, value=e.cantidad)
        c.number_format = number_format
        c.border = thin_border
        row += 1

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 14

    ws3 = wb.create_sheet("Consolidado por Colegio")

    ws3.merge_cells("A1:F1")
    ws3["A1"] = "CONSOLIDADO POR COLEGIO"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="4F46E5")
    ws3["A1"].alignment = Alignment(horizontal="center")

    consolidado = {}
    all_votos = Voto.query.all()
    for v in all_votos:
        key = (v.colegio_id, v.colegio.nombre if v.colegio else "N/A")
        if key not in consolidado:
            consolidado[key] = {}
        if v.partido_nombre not in consolidado[key]:
            consolidado[key][v.partido_nombre] = 0
        consolidado[key][v.partido_nombre] += v.votos

    if consolidado:
        all_partidos = sorted(set(pnombre for datos in consolidado.values() for pnombre in datos.keys()))
        headers3 = ["Colegio"] + all_partidos + ["Total"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            cell.alignment = header_alignment
            cell.border = thin_border

        row = 4
        for (cid, cnombre), partidos in sorted(consolidado.items()):
            ws3.cell(row=row, column=1, value=cnombre).border = thin_border
            total_row = 0
            for i, pnombre in enumerate(all_partidos):
                val = partidos.get(pnombre, 0)
                c = ws3.cell(row=row, column=i + 2, value=val)
                c.number_format = number_format
                c.border = thin_border
                total_row += val
            c = ws3.cell(row=row, column=len(all_partidos) + 2, value=total_row)
            c.number_format = number_format
            c.font = Font(bold=True)
            c.border = thin_border
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"votos_{peru_now().strftime('%Y%m%d_%H%M')}.xlsx",
    )


@votos_bp.route("/api/exportar/pdf")
@login_required
def exportar_pdf():
    from fpdf import FPDF
    from flask import send_file
    import io

    colegio_id = request.args.get("colegio_id", type=int)

    query_votos = Voto.query
    query_especial = VotoEspecial.query

    if colegio_id:
        query_votos = query_votos.filter_by(colegio_id=colegio_id)
        query_especial = query_especial.filter_by(colegio_id=colegio_id)

    votos = query_votos.order_by(Voto.cargo, Voto.partido_nombre).all()
    especiales = query_especial.order_by(VotoEspecial.cargo).all()

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(79, 70, 229)
    pdf.cell(0, 12, "REPORTE DE VOTOS - ELECCIONES 2026", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(107, 114, 128)
    pdf.cell(0, 6, f"Generado: {peru_now().strftime('%d/%m/%Y %H:%M')}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(6)

    from models.voto import CARGOS
    cargos_map = {c["id"]: c["nombre"] for c in CARGOS}

    for cargo_id, cargo_nombre in cargos_map.items():
        cargo_votos = [v for v in votos if v.cargo == cargo_id]
        cargo_especiales = [e for e in especiales if e.cargo == cargo_id]

        if not cargo_votos and not cargo_especiales:
            continue

        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(79, 70, 229)
        pdf.cell(0, 8, cargo_nombre, new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(79, 70, 229)
        pdf.line(10, pdf.get_y(), 287, pdf.get_y())
        pdf.ln(3)

        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(79, 70, 229)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(80, 7, "Partido", border=1, fill=True, align="C")
        pdf.cell(25, 7, "Sigla", border=1, fill=True, align="C")
        pdf.cell(30, 7, "Votos", border=1, fill=True, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(0, 0, 0)
        total_cargo = 0
        for v in cargo_votos:
            pdf.cell(80, 6, v.partido_nombre[:40], border=1)
            pdf.cell(25, 6, v.partido_sigla, border=1, align="C")
            pdf.cell(30, 6, str(v.votos), border=1, align="C")
            pdf.ln()
            total_cargo += v.votos

        if cargo_especiales:
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(80, 6, "VOTOS ESPECIALES", border=1, fill=False)
            pdf.cell(25, 6, "", border=1)
            pdf.cell(30, 6, "", border=1)
            pdf.ln()
            pdf.set_font("Helvetica", "", 8)
            for e in cargo_especiales:
                pdf.cell(80, 6, f"  {e.tipo}", border=1)
                pdf.cell(25, 6, "", border=1, align="C")
                pdf.cell(30, 6, str(e.cantidad), border=1, align="C")
                pdf.ln()
                total_cargo += e.cantidad

        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(105, 7, f"TOTAL {cargo_nombre.upper()}", border=1, fill=True)
        pdf.cell(30, 7, str(total_cargo), border=1, fill=True, align="C")
        pdf.ln(10)

    total_votos_partidos = sum(v.votos for v in votos)
    total_votos_especiales = sum(e.cantidad for e in especiales)

    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(79, 70, 229)
    pdf.cell(0, 10, "RESUMEN GENERAL", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(120, 8, "Total Votos Partidos:", border=1)
    pdf.cell(50, 8, str(total_votos_partidos), border=1, align="C")
    pdf.ln()
    pdf.cell(120, 8, "Total Votos Especiales:", border=1)
    pdf.cell(50, 8, str(total_votos_especiales), border=1, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_fill_color(79, 70, 229)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(120, 9, "TOTAL GENERAL:", border=1, fill=True)
    pdf.cell(50, 9, str(total_votos_partidos + total_votos_especiales), border=1, fill=True, align="C")
    pdf.ln()

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"reporte_votos_{peru_now().strftime('%Y%m%d_%H%M')}.pdf",
    )
